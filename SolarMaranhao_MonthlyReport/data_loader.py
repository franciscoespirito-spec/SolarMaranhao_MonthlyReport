"""
Extração de dados do Excel de geração solar.
Lê dados diários e mensais das abas de cada usina e do Sumário.
"""
import warnings
from datetime import datetime, date
from calendar import monthrange

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string

from config import (
    PLANTS, PLANT_IDS, EXCEL_PATH,
    EXCEL_YEAR_COLUMNS, MONTH_OFFSET_2026,
    SUMARIO_MONTHLY_COLS, SUMARIO_MONTHLY_START_ROW, SUMARIO_MONTHLY_END_ROW,
)

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def _safe_float(val):
    """Converte valor para float, tratando erros do Excel."""
    if val is None:
        return 0.0
    if isinstance(val, str):
        if val.startswith("#") or val == "-":
            return 0.0
        try:
            return float(val)
        except ValueError:
            return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _col_idx(col_letter):
    """Converte letra de coluna para índice base-1."""
    return column_index_from_string(col_letter)


def load_workbook(excel_path=None):
    """Carrega o workbook com valores calculados."""
    path = excel_path or EXCEL_PATH
    return openpyxl.load_workbook(path, data_only=True)


def load_monthly_totals_from_sheets(wb, year=2026):
    """
    Lê totais mensais de cada aba de usina (linha 5 da seção de resumo do ano).
    Retorna DataFrame com colunas: month (1-12), spe1..spe6.
    Fonte mais confiável que o Sumario para meses recentes.
    """
    year_cfg = EXCEL_YEAR_COLUMNS.get(year)
    if not year_cfg:
        return pd.DataFrame()

    # Mapeamento de mês para coluna (2026: CI=Jan, CJ=Feb, etc.)
    if year == 2026:
        month_col_map = MONTH_OFFSET_2026
    else:
        # Para outros anos, calcular a partir de month_cols_start
        start_letter = year_cfg["month_cols_start"]
        start_idx = _col_idx(start_letter)
        month_col_map = {}
        for m in range(1, 13):
            col_idx = start_idx + (m - 1)
            month_col_map[m] = openpyxl.utils.get_column_letter(col_idx)

    data = []
    for month_num, col_letter in month_col_map.items():
        col_i = _col_idx(col_letter)
        row_data = {"month": month_num}

        for plant_id in PLANT_IDS:
            plant = PLANTS[plant_id]
            ws = wb[plant["sheet_name"]]
            # Linha 5 = "Somado da tabela (kWh)" na seção de resumo
            val = ws.cell(row=5, column=col_i).value
            row_data[plant_id] = _safe_float(val)

        data.append(row_data)

    df = pd.DataFrame(data)
    df["total"] = sum(df[pid] for pid in PLANT_IDS)
    return df


def load_monthly_totals_from_sumario(wb):
    """
    Lê totais mensais do Sumario (AS:AX, linhas 5-52).
    Cobre todos os anos disponíveis.
    Retorna DataFrame com: date, spe1..spe6, total.
    """
    ws = wb["Sumario"]
    date_col = _col_idx(SUMARIO_MONTHLY_COLS["date_col"])

    data = []
    for row in range(SUMARIO_MONTHLY_START_ROW, SUMARIO_MONTHLY_END_ROW + 1):
        dt = ws.cell(row=row, column=date_col).value
        if dt is None:
            continue
        if isinstance(dt, datetime):
            dt = dt.date()
        elif isinstance(dt, str):
            try:
                dt = datetime.strptime(dt, "%Y-%m-%d").date()
            except ValueError:
                continue

        row_data = {"date": dt, "year": dt.year, "month": dt.month}
        for plant_id in PLANT_IDS:
            col = _col_idx(SUMARIO_MONTHLY_COLS[plant_id])
            row_data[plant_id] = _safe_float(ws.cell(row=row, column=col).value)

        row_data["total"] = sum(row_data[pid] for pid in PLANT_IDS)
        data.append(row_data)

    return pd.DataFrame(data)


def load_daily_data(wb, year, month):
    """
    Lê dados diários de cada usina para um mês específico.
    Retorna DataFrame com: date, spe1..spe6.
    """
    year_cfg = EXCEL_YEAR_COLUMNS.get(year)
    if not year_cfg:
        return pd.DataFrame()

    date_col_idx = _col_idx(year_cfg["date_col"])
    gen_col_idx = _col_idx(year_cfg["gen_col"])

    _, days_in_month = monthrange(year, month)
    target_start = date(year, month, 1)
    target_end = date(year, month, days_in_month)

    # Coletar datas e geração de cada usina
    plant_daily = {}
    for plant_id in PLANT_IDS:
        plant = PLANTS[plant_id]
        ws = wb[plant["sheet_name"]]
        daily = {}

        for row in range(4, ws.max_row + 1):
            dt = ws.cell(row=row, column=date_col_idx).value
            if dt is None:
                continue
            if isinstance(dt, datetime):
                dt = dt.date()
            elif not isinstance(dt, date):
                continue

            if target_start <= dt <= target_end:
                gen = _safe_float(ws.cell(row=row, column=gen_col_idx).value)
                daily[dt] = gen

        plant_daily[plant_id] = daily

    # Montar DataFrame
    all_dates = sorted(set().union(*(d.keys() for d in plant_daily.values())))
    rows = []
    for dt in all_dates:
        row = {"date": dt}
        for plant_id in PLANT_IDS:
            row[plant_id] = plant_daily[plant_id].get(dt, 0.0)
        rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty:
        df["total"] = sum(df[pid] for pid in PLANT_IDS)
    return df


def load_all_monthly_totals(wb, target_year=2026):
    """
    Combina dados do Sumario (anos antigos) com dados das abas individuais
    (ano corrente) para ter o histórico completo.
    Retorna DataFrame com: year, month, spe1..spe6, total.
    """
    # Sumario para dados históricos
    sumario_df = load_monthly_totals_from_sumario(wb)

    # Abas individuais para o ano corrente (mais confiável)
    sheets_df = load_monthly_totals_from_sheets(wb, year=target_year)
    sheets_df["year"] = target_year

    if sumario_df.empty:
        return sheets_df

    # Remover dados do ano corrente do Sumario (substituir pelos das abas)
    sumario_old = sumario_df[sumario_df["year"] != target_year].copy()

    # Combinar
    combined = pd.concat([sumario_old, sheets_df], ignore_index=True)
    combined = combined.sort_values(["year", "month"]).reset_index(drop=True)

    return combined


def load_sumario_kpis(wb):
    """
    Lê os KPIs live do Sumario (linhas 5-11, colunas E-K).
    Retorna dict com as métricas atuais.
    """
    ws = wb["Sumario"]

    kpis = {}
    labels = {
        5: "geracao_acumulada_mes",
        6: "geracao_acumulada_ano",
        7: "media_geracao_mes",
        8: "media_diaria_ano",
        9: "projecao_mensal",
        10: "geracao_maxima_ano",
        11: "media_vs_maxima",
    }

    for row_num, key in labels.items():
        kpis[key] = {}
        for plant_id in PLANT_IDS:
            col = _col_idx("F") + PLANT_IDS.index(plant_id)
            kpis[key][plant_id] = _safe_float(ws.cell(row=row_num, column=col).value)
        # Total na coluna K
        kpis[key]["total"] = _safe_float(ws.cell(row=row_num, column=_col_idx("K")).value)

    return kpis
